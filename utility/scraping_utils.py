import json
import os
import random
import re
import sys
import time
from collections import Counter
from datetime import datetime

import pandas as pd
from bs4 import BeautifulSoup
from fuzzywuzzy import fuzz, process
from openpyxl import Workbook, load_workbook


def get_current_time_omnizia_format():
    current_datetime = datetime.utcnow()
    formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S.%f+00")
    return formatted_datetime


def get_array_from_xlsx_files_columns(xlsx_file_path, column_name):
    try:
        df = pd.read_excel(xlsx_file_path)
        if column_name not in df.columns:
            return None
        column_values = df[column_name].values
        return column_values
    except Exception as e:
        print(f"Error occurred while reading the file: {e}", file=sys.stderr)
        return None


def null_check(value):
    if value is None:
        value = ''
    return str(value.strip())


def load_csv_into_df(csv_file_path):
    try:
        df = pd.read_csv(csv_file_path, low_memory=False)
        return df
    except Exception as e:
        print(f'Not able to load {csv_file_path} file. Exception: {e}', file=sys.stderr)
        return None


def write_to_excel(file_name, sheet_name, df):
    try:
        # Load existing workbook if it exists, otherwise create a new one
        try:
            wb = load_workbook(file_name)
        except FileNotFoundError:
            wb = Workbook()

        # Check if the sheet already exists, if not create a new one
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Find the next empty row in the sheet
            next_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title=sheet_name)
            next_row = 1

        # Convert DataFrame to list of lists and write to worksheet
        data = df.values.tolist()
        for row in data:
            ws.append(row)

        # Save the workbook
        wb.save(file_name)
        print(f"Data written to {file_name} under sheet '{sheet_name}'")
    except Exception as e:
        print("Error occurred:", e)


def write_object_to_excel(file_name, sheet_name, obj):
    try:
        # Load existing workbook if it exists, otherwise create a new one
        try:
            wb = load_workbook(file_name)
        except FileNotFoundError:
            wb = Workbook()

        # Check if the sheet already exists, if not create a new one
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            next_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title=sheet_name)
            next_row = 1

        # Extract the property names (header names) from the object
        headers = list(obj.__dict__.keys())

        # Write headers to the first row of the worksheet
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Write values to the next row
        for col, header in enumerate(headers, start=1):
            value = getattr(obj, header)
            ws.cell(row=next_row, column=col, value=value)

        # Save the workbook
        wb.save(file_name)
        print(f"Object written to {file_name} under sheet '{sheet_name}'")
    except Exception as e:
        print("Error occurred:", e)


def write_objects_to_excel(file_name, sheet_name, objects_list):
    try:
        # Load existing workbook if it exists, otherwise create a new one
        try:
            wb = load_workbook(file_name)
        except FileNotFoundError:
            wb = Workbook()

        # Check if the sheet already exists, if not create a new one
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            next_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title=sheet_name)
            next_row = 1

        # If objects_list is empty, just save the workbook and return
        if not objects_list:
            wb.save(file_name)
            print(f"No data to write. {file_name} is saved without changes.")
            return

        # Extract headers from the objects in objects_list
        headers = set()
        for obj in objects_list:
            if isinstance(obj, dict):
                headers.update(obj.keys())
            else:
                headers.update(obj.__dict__.keys())

        # Write headers to the first row of the worksheet if they are not already written
        if next_row == 1 and all(ws.cell(row=1, column=col).value is None for col in range(1, len(headers) + 1)):
            for col, header in enumerate(headers, start=1):
                ws.cell(row=next_row, column=col, value=header)
            next_row += 1

        # Check for new headers and write them if needed
        for col, header in enumerate(headers, start=1):
            if ws.cell(row=1, column=col).value is None:
                ws.cell(row=1, column=col, value=header)

        # Write values for each object in the list
        for obj in objects_list:
            for col, header in enumerate(headers, start=1):
                if isinstance(obj, dict):
                    value = obj.get(header, "")
                else:
                    value = getattr(obj, header, "")
                ws.cell(row=next_row, column=col, value=value)
            next_row += 1

        # Save the workbook
        wb.save(file_name)
        print(f"Objects written to {file_name} under sheet '{sheet_name}'")
    except Exception as e:
        print("Error occurred:", e)


def write_objects_to_excel2(file_name, sheet_name, objects_list):
    try:
        # Load existing workbook if it exists, otherwise create a new one
        try:
            wb = load_workbook(file_name)
        except FileNotFoundError:
            wb = Workbook()

        # Check if the sheet already exists, if not create a new one
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            next_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title=sheet_name)
            next_row = 1

        # If objects_list is empty, just save the workbook and return
        if not objects_list:
            wb.save(file_name)
            print(f"No data to write. {file_name} is saved without changes.")
            return

        # Extract the property names (header names) from the first object
        headers = list(objects_list[0].__dict__.keys())

        # Write headers to the first row of the worksheet if they are not already written
        if next_row == 1 and all(ws.cell(row=1, column=col).value is None for col in range(1, len(headers) + 1)):
            for col, header in enumerate(headers, start=1):
                ws.cell(row=next_row, column=col, value=header)
            next_row += 1

        # Check for new headers and write them if needed
        for col, header in enumerate(headers, start=1):
            if ws.cell(row=1, column=col).value is None:
                ws.cell(row=1, column=col, value=header)

        # Write values for each object in the list
        for obj in objects_list:
            for col, header in enumerate(headers, start=1):
                value = getattr(obj, header)
                ws.cell(row=next_row, column=col, value=value)
            next_row += 1

        # Save the workbook
        wb.save(file_name)
        print(f"Objects written to {file_name} under sheet '{sheet_name}'")
    except Exception as e:
        print("Error occurred:", e)


def load_xlsx_into_df(file_path, sheet_name):
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        return df
    except Exception as e:
        print(f'Not able to load {file_path} file. Exception: {e}', file=sys.stderr)
        return None





def divide_with_four_decimal_places(divisor, divider):
    result_with_four_decimals = 0
    try:
        result = divisor / divider
        result_with_four_decimals = "{:.4f}".format(result)
    except Exception as e:
        print(
            f'Exception occurred when divide {divisor} by {divider}. Exception: {e}', file=sys.stderr)
    return float(result_with_four_decimals)


def generate_df_to_csv(df, file_name=None):
    try:
        if file_name is None or len(str(file_name)) == 0:
            file_name = 'output_file.csv'
        df.to_csv(file_name, index=False)
        print(f'{file_name} file is generated.', file=sys.stdout)
    except Exception as e:
        print(f'Unable to generate {file_name} file. Exception: {e}', file=sys.stderr)


def csv_to_xlsx(csv_file_path, xlsx_file_path='output.xlsx'):
    """
    csv to exel converter

    :param csv_file_path:
    :param xlsx_file_path:
    :return:
    """
    try:
        df = pd.read_csv(csv_file_path)
        df.to_excel(xlsx_file_path, index=False)
        print_output(f'CSV file "{csv_file_path}" converted to XLSX file "{xlsx_file_path}"')
    except Exception as e:
        print_error(f'Exception occurred when converting csv to xlsx. Message: {e}')


def xlsx_to_csv(xlsx_filename, csv_filename='data/output.csv', sheet_name=None, column_to_read=None):
    """
    Exel file to csv converter

    :param xlsx_filename: This is the input filename
    :param csv_filename: This is the output filename
    :param sheet_name: Sheet name. Default is none which take active sheet.
    :param column_to_read:
    :return:
    """
    if xlsx_filename is None:
        raise Exception('Input filename is required')
    try:
        if sheet_name is None:
            df = pd.read_excel(xlsx_filename, usecols=column_to_read)
        else:
            df = pd.read_excel(xlsx_filename, sheet_name=sheet_name, usecols=column_to_read)

        df.to_csv(csv_filename, index=False)
        print_output(f"Conversion successful: '{xlsx_filename}' converted to '{csv_filename}'")
    except Exception as e:
        print_error(f"Error during conversion: {e}")


def print_output(message):
    print(f'{message}', file=sys.stdout)


def print_error(message):
    print(f'{message}', file=sys.stderr)


def duplicate_remove(csv_file, column):
    """
    This function can remove duplicate rows from a specific csv file.

    :param csv_file: The csv filename. Can be relative path.
    :param column: The column name of the csv file which is the unique identifier
    :return: This function return nothing
    """
    data = pd.read_csv(csv_file)
    column_to_check = column

    # Find and remove duplicates while keeping only the first occurrence
    data_unique = data.drop_duplicates(subset=column_to_check, keep="first")

    # Save the cleaned DataFrame back to a CSV file
    cleaned_csv_file = "data/cleaned_file.csv"
    data_unique.to_csv(cleaned_csv_file, index=False)

    print_output("Duplicates removed and cleaned CSV file saved successfully.")


def save_object_to_csv(obj, file_name=None, keep_duplicate=False):
    """
    This function is responsible to save data into csv file.

    :param obj: A python object. It will be the row of csv file
    :param file_name: This will be the csv file name. This can be relative path
    :return: This function return nothing
    """
    if file_name is None:
        file_name = 'data/example.csv'

    df = pd.DataFrame([obj])
    old_df = load_csv_into_df(file_name)
    modified_df = pd.concat([old_df, df], ignore_index=True)
    if keep_duplicate:
        modified_df = modified_df.drop_duplicates()
    modified_df.to_csv(file_name, index=False)
    print_output("Data saved")


def save_list_of_object_to_csv(list, file_name=None, keep_duplicate=False):
    """
    This function is responsible to save list of object into a specific csv file

    :param list: Receive a list of object
    :param file_name: This will be the csv filename. This can be relative path.
    :return: This function return nothing
    """
    try:
        if file_name is None:
            file_name = 'data/example.csv'

        df = pd.DataFrame(list)
        old_df = load_csv_into_df(file_name)
        modified_df = pd.concat([old_df, df], ignore_index=True)
        if keep_duplicate:
            modified_df = modified_df.drop_duplicates()
        modified_df.to_csv(file_name, index=False)

        return True, 'Saved'
    except Exception as e:
        return False, e


def csv_to_csv(input_file_path, columns=None, output_file_path='output.csv'):
    if not isinstance(columns, list) or columns is None:
        columns = []

    df_source = pd.read_csv(input_file_path)
    if len(columns) > 0:
        df_selected_columns = df_source[columns]
        df_selected_columns.to_csv(output_file_path, index=False)
    else:
        df_source.to_csv(output_file_path)


def html_to_text(html_text):
    soup = BeautifulSoup(html_text, 'html.parser')
    plain_text = soup.get_text()
    return plain_text


def string_similarity(str1, str2, case_sensitive=True):
    """
    Compare two strings and return a similarity ratio.

    :param str1: The first string for comparison.
    :param str2: The second string for comparison.
    :param case_sensitive: A boolean parameter to specify whether the comparison is case-sensitive (default is True).
    :return: A similarity ratio from 0 to 100.
    """
    if case_sensitive:
        ratio = fuzz.ratio(str1, str2)
    else:
        ratio = fuzz.ratio(str1.lower(), str2.lower())
    return ratio


def remove_trailing_string(input_string, trailing_string):
    try:
        # Use rstrip to remove the trailing string from the right side
        result_string = input_string.rstrip(trailing_string)

        return result_string
    except:
        return input_string



def normalize_whitespace(text):
    # Use regular expression to replace multiple whitespaces with a single whitespace
    normalized_text = re.sub(r'\s+', ' ', text)
    return normalized_text.strip()


def generate_random_number(number_of_digits):
    if number_of_digits <= 0:
        raise ValueError("Number of digits should be greater than zero.")

    # Generate a random number with the specified number of digits
    lower_bound = 10 ** (number_of_digits - 1)
    upper_bound = 10 ** number_of_digits - 1
    random_number = random.randint(lower_bound, upper_bound)

    return random_number


def generate_random_number_with_possible_leading_zero(number_of_digits):
    if number_of_digits <= 0:
        raise ValueError("Number of digits should be greater than zero.")

    # Generate a list of random digits
    random_digits = [str(random.randint(0, 9)) for _ in range(number_of_digits)]

    # Join the digits together to form the random number
    random_number = str(''.join(random_digits))

    return random_number


def generate_random_omnizia_id(country_iso2):
    vqp = 'VQP'
    iso = str(country_iso2).strip().upper()
    num = generate_random_number_with_possible_leading_zero(5)

    new_omnizia_id = vqp + iso + '000' + num
    return new_omnizia_id


def save_list_of_str_into_file(FILE_PATH, list):
    try:
        with open(FILE_PATH, 'w') as file:
            for item in list:
                if len(item.strip()) == 0:
                    continue
                file.write(item + '\n')
        return True, f'Data saved in {FILE_PATH}'

    except Exception as e:
        return False, f'Not able to save in {FILE_PATH}. Error: {e}'


def get_list_of_str_from_file(FILE_PATH):
    try:
        loaded_array = []
        with open(FILE_PATH, 'r') as file:
            for line in file:
                item = line.strip()
                loaded_array.append(item)
        return loaded_array
    except Exception as e:
        return f'Not able to retrieved from {FILE_PATH}. Error: {e}'


def is_numeric_integer(s):
    try:
        number = int(s)
        if isinstance(number, int):
            num_digits = len(str(abs(number)))
            return True, num_digits
    except ValueError:
        pass
    return False, 0


def get_string_from_file(FILE_PATH):
    try:
        with open(FILE_PATH, 'r') as file:
            return file.readline().strip()
    except Exception as e:
        return f'Not able to retrieved from {FILE_PATH}. Error: {e}'


def remove_whitespaces(string_value):
    return str(string_value).strip()


def to_lower(string):
    return get_empty_if_null(string).lower()


def to_upper(string):
    return get_empty_if_null(string).upper()



def duplicate_remove_from_csv(input_file_path, columns_to_check_duplicates, output_file_path):
    try:
        df = load_csv_into_df(input_file_path)
        # Drop duplicates based on the specified columns
        df_unique = df.drop_duplicates(subset=columns_to_check_duplicates, keep='first')
        # Save the DataFrame with unique entries to a new CSV file
        df_unique.to_csv(output_file_path, index=False)
        return True, 'Success'
    except Exception as e:
        return False, f'Failed. Error: {e}'


def get_unique_list(list_having_duplicate):
    # Remove duplicates by converting the array to a set
    unique_set = set(list_having_duplicate)
    # Convert the set back to a list if needed
    unique_list = list(unique_set)
    return unique_list


def print_exception(e):
    print(f'Exception: {str(e)}', file=sys.stderr)



def is_json(data):
    try:
        json.loads(data)
        return True
    except:
        return False


def str_to_json(json_string):
    try:
        return json.loads(json_string)
    except:
        return None


def char_count(input_string: str, char_to_be_count: str, ignore_case=False):
    if ignore_case:
        input_string = input_string.lower()
        char_to_be_count = char_to_be_count.lower()
    return input_string.count(char_to_be_count)


def get_pandas():
    return pd


def find_best_matches(main_string, string_list, threshold=70):
    """
    Find best matches between a main string and a list of strings.

    Parameters:
    - main_string: The string to compare against.
    - string_list: List of strings to search for matches.
    - threshold: The similarity threshold (default is 70).

    Returns:
    - List of best-matching strings, potentially including multiple matches.
    """
    # Get matches with similarity scores
    matches = process.extract(main_string, string_list, limit=len(string_list))

    # Find the best match
    best_match_score = matches[0][1]
    best_matches = [match[0] for match in matches if match[1] == best_match_score and match[1] >= threshold]

    return best_matches


def list_files(directory):
    file_paths = []

    for root, dirs, files in os.walk(directory):
        for file in files:
            # Get the relative path by joining the root and file
            relative_path = os.path.relpath(os.path.join(root, file), directory)
            file_paths.append(relative_path)

    return file_paths


def find_most_common_string(list_of_string: []):
    try:
        counter = Counter(list_of_string)
        # Find the most common element and its count
        most_common_entry = counter.most_common(1)
        entry, count = most_common_entry[0]
        return entry
    except:
        return None


def rename_csv_file(original_file_path, new_file_path):
    df = pd.read_csv(original_file_path)
    df.to_csv(new_file_path, index=False)


def replace_special_characters(input_string):
    # Define a regular expression to match special characters
    special_char_pattern = re.compile(r'[^\w.-]')
    # Replace special characters with underscores
    output_string = re.sub(special_char_pattern, '_', input_string)
    return output_string


def combine_csv_files(directory_path, output_file_name='combined_file.csv'):
    # Get a list of all CSV files in the specified directory
    csv_files = [file for file in os.listdir(directory_path) if file.endswith('.csv')]
    # Check if there are any CSV files in the directory
    if not csv_files:
        print("No CSV files found in the specified directory.")
        return
    # Initialize an empty DataFrame to store the combined data
    combined_data = pd.DataFrame()
    # Iterate through each CSV file and append its data to the combined DataFrame
    for csv_file in csv_files:
        file_path = os.path.join(directory_path, csv_file)
        df = pd.read_csv(file_path)
        combined_data = combined_data._append(df, ignore_index=True)

    # Write the combined DataFrame to a new CSV file
    output_file_path = os.path.join(directory_path, output_file_name)
    combined_data.to_csv(output_file_path, index=False)

    print(f"Combined CSV file saved to: {output_file_path}")


def remove_duplicates(input_df):
    # Drop duplicates based on all columns
    output_df = input_df.drop_duplicates()
    return output_df


def get_current_time_millis():
    current_time_seconds = time.time()
    current_time_millis = int(current_time_seconds * 1000)
    return current_time_millis


def extract_links_from_text(text):
    # Regular expression pattern to match URLs
    url_pattern = r'https?://[^\s]+'

    # Find all URLs in the text
    links = re.findall(url_pattern, text)

    return links


def extract_number_from_url(url):
    # Regular expression to find decimal integers with more than 6 digits in the URL
    pattern = r'\b\d{7,}\b'

    # Search for the pattern in the URL
    match = re.search(pattern, url)

    if match:
        # Convert the matched string to an integer
        number = int(match.group())
        return number
    else:
        return ''


def remove_special_characters(special_chars, text):
    # Iterate over each character in the text
    cleaned_text = ''.join(char for char in text if char not in special_chars)
    return cleaned_text
